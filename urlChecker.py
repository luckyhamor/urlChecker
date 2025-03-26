import sys
import re
import socket
import requests
import time
import argparse
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

EXPECTED_DOMAIN = "https://example.com"
VIRUSTOTAL_API_KEY = "VirusTotal_APIKEY"
VIRUSTOTAL_URL = "https://www.virustotal.com/api/v3/urls"

SUSPICIOUS_PATTERNS = [
    r"https?://[0-9]+\.[0-9]+\.[0-9]+\.[0-9]+",
    r"https?://.*@",
    r"https?://.*\?.*=",
    r"https?://.*#.*",
    r"https?://.*\.(zip|exe|apk|bat|cmd|sh|php)$"
]

def load_urls(file_path):
    urls = []
    try:
        with open(file_path, "r") as file:
            for line in file:
                line = line.strip()
                if line:
                    parts = line.split("\t") if "\t" in line else line.rsplit(" ", 1)
                    if len(parts) == 2 and parts[0] in ["GET", "POST", "PUT", "DELETE"]:
                        method, path = parts
                    elif len(parts) == 2 and parts[1] in ["GET", "POST", "PUT", "DELETE"]:
                        path, method = parts
                    else:
                        method, path = "GET", line
                    urls.append((method.upper(), path))
        return urls
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)

def is_rogue_url(method, url):
    full_url = EXPECTED_DOMAIN + url if url.startswith("/") else url
    print(f"Scanning URL: {full_url} (Method: {method})")
    for pattern in SUSPICIOUS_PATTERNS:
        if re.search(pattern, full_url):
            return full_url, "Malicious"
    try:
        domain = full_url.split("/")[2]
        socket.gethostbyname(domain)
    except (IndexError, socket.gaierror):
        return full_url, "Malicious"
    return full_url, "Malicious" if check_virustotal(full_url) else "Safe"

def check_virustotal(url):
    headers = {"x-apikey": VIRUSTOTAL_API_KEY}
    response = requests.post(VIRUSTOTAL_URL, headers=headers, json={"url": url})
    if response.status_code == 200:
        analysis_id = response.json().get("data", {}).get("id")
        if analysis_id:
            return get_virustotal_analysis(analysis_id)
    return False

def get_virustotal_analysis(analysis_id):
    headers = {"x-apikey": VIRUSTOTAL_API_KEY}
    analysis_url = f"https://www.virustotal.com/api/v3/analyses/{analysis_id}"
    for _ in range(5):
        response = requests.get(analysis_url, headers=headers)
        if response.status_code == 200:
            stats = response.json().get("data", {}).get("attributes", {}).get("stats", {})
            return stats.get("malicious", 0) > 0
        time.sleep(5)
    return False

def find_rogue_urls(url_list):
    results = [is_rogue_url(method, url) for method, url in url_list]
    return results

def save_results(results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["URL", "Status"])
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for url, status in results:
        row = [url, status]
        ws.append(row)
        if status == "Malicious":
            ws[f"A{ws.max_row}"].fill = red_fill
            ws[f"B{ws.max_row}"].fill = red_fill
    
    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description="Scan URLs for malicious patterns.")
    parser.add_argument("file", nargs="?", help="File containing list of URLs")
    parser.add_argument("-u", "--url", help="Single URL to scan")
    parser.add_argument("-o", "--output", help="Output file to save results")
    args = parser.parse_args()

    urls = []
    if args.url:
        urls.append(("GET", args.url))
    elif args.file:
        urls = load_urls(args.file)
    else:
        print("Usage: python3 urlCheck.py [-u URL] [file] [-o OUTPUT]")
        sys.exit(1)
    
    results = find_rogue_urls(urls)
    
    print("\nScan Results:")
    print("{:<60} {:<10}".format("URL", "Status"))
    print("-" * 70)
    for url, status in results:
        print(f"{url:<60} {status}")
    
    if args.output:
        save_results(results, args.output)
        print(f"\nResults saved to {args.output}")

    total_scanned = len(results)
    safe_count = sum(1 for _, status in results if status == "Safe")
    malicious_count = total_scanned - safe_count
    print(f"\nSummary: {safe_count}/{total_scanned} are Safe, {malicious_count}/{total_scanned} are Malicious")

if __name__ == "__main__":
    main()
