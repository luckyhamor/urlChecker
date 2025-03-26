import sys
import socket
import requests
import time
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill

EXPECTED_DOMAIN = "https://example.com"
VIRUSTOTAL_API_KEY = "APIKEY"
VIRUSTOTAL_URL = "https://www.virustotal.com/api/v3/urls"

def load_urls(file_path):
    """Load URLs from a file, clean them, and extract methods."""
    urls = []
    try:
        with open(file_path, "r") as file:
            for line in file:
                line = line.strip().replace('"', '').replace("\t", " ")  # Remove quotes and tabs
                if not line:
                    continue  # Skip empty lines

                parts = line.rsplit(" ", 1)  # Split by last space to separate method
                if len(parts) == 2 and parts[1].upper() in ["GET", "POST", "PUT", "DELETE"]:
                    path, method = parts
                else:
                    path, method = line, "GET"  # Default to GET if method is missing

                path = path.strip()
                full_url = EXPECTED_DOMAIN + path if path.startswith("/") else path
                urls.append((method.upper(), full_url))
        return urls
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)

def check_virustotal(url):
    """Check URL with VirusTotal."""
    headers = {"x-apikey": VIRUSTOTAL_API_KEY}
    response = requests.post(VIRUSTOTAL_URL, headers=headers, json={"url": url})
    if response.status_code == 200:
        analysis_id = response.json().get("data", {}).get("id")
        if analysis_id:
            return get_virustotal_analysis(analysis_id)
    return "Unknown"

def get_virustotal_analysis(analysis_id):
    """Retrieve VirusTotal scan results."""
    headers = {"x-apikey": VIRUSTOTAL_API_KEY}
    analysis_url = f"https://www.virustotal.com/api/v3/analyses/{analysis_id}"
    for _ in range(5):  # Retry up to 5 times
        response = requests.get(analysis_url, headers=headers)
        if response.status_code == 200:
            stats = response.json().get("data", {}).get("attributes", {}).get("stats", {})
            malicious_count = stats.get("malicious", 0)
            return "Malicious" if malicious_count > 0 else "Safe"
        time.sleep(5)
    return "Unknown"

def scan_urls(url_list):
    """Scan URLs and determine their status."""
    results = []
    for method, url in url_list:
        print(f"Scanning URL: {url} (Method: {method})")
        status = check_virustotal(url)
        results.append((url, method, status))
    return results

def save_results(results, output_file):
    """Save results to an Excel file and highlight malicious rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(["URL", "Method", "Status"])

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for url, method, status in results:
        row = [url, method, status]
        ws.append(row)
        if status == "Malicious":
            for col in range(1, 4):  # Highlight all columns in the row
                ws[f"{chr(64+col)}{ws.max_row}"].fill = red_fill

    wb.save(output_file)

def main():
    parser = argparse.ArgumentParser(description="Scan URLs for threats using VirusTotal.")
    parser.add_argument("-f", "--file", help="File containing list of URLs")
    parser.add_argument("-o", "--output", help="Output Excel file", required=True)
    args = parser.parse_args()

    if not args.file:
        print("Error: You must provide a file with URLs (-f <filename>)")
        sys.exit(1)

    urls = load_urls(args.file)
    results = scan_urls(urls)

    # Print summary
    safe_count = sum(1 for _, _, status in results if status == "Safe")
    malicious_count = sum(1 for _, _, status in results if status == "Malicious")
    total_scanned = len(results)

    print("\nSummary:")
    print(f"✅ Safe: {safe_count}/{total_scanned}")
    print(f"❌ Malicious: {malicious_count}/{total_scanned}")

    save_results(results, args.output)
    print(f"\nResults saved to {args.output}")

if __name__ == "__main__":
    main()

